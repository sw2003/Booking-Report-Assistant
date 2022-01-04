from openpyxl import load_workbook
from openpyxl.workbook import workbook
from flask import Blueprint, app, render_template, request, current_app, send_from_directory
import os

write_app = Blueprint("write_app", __name__, static_folder="static", template_folder="template") 


@write_app.route("/fill", methods = ["POST"])
def fill_data():
    if len(os.listdir("static/Sheets")) >= 1:
        for i in os.listdir("static/Sheets"):
            os.remove("static/Sheets/{}".format(i))

    if request.files:
        read_file = request.files['read_file']
        write_file = request.files['write_file']
        
        for file in [read_file, write_file]:
            file.save(os.path.join(current_app.config["SHEET_UPLOAD"], file.filename))

        main_workbook = load_workbook(filename = 'static/Sheets/{}'.format(write_file.filename))
        main_ws = main_workbook.active

        read_workbook = load_workbook(filename = 'static/Sheets/{}'.format(read_file.filename))
        read_ws = read_workbook.active

        pointer = 5

        shipping_lane_names = ['CPNW', 'CENX','MPNW', 'OPNW', 'EPNW', 'AWE5', 'GEX1', 'GEX2']
        shipping_lane_code = ""

        for j in range(5, main_ws.max_row-6):
            name = main_ws['A{}'.format(j)].value
            color_of_tab = main_ws['A{}'.format(j)].fill.fgColor
            
            if name != "BLANK VOYAGE" and name != None and name != "TBN" and name != "SUB TOTAL":
                if name in shipping_lane_names or color_of_tab.tint == 0:
                    
                    name = name.strip()
                    if name in shipping_lane_names:
                        shipping_lane_code = name
                    else:
                        number_code = name.strip().split(" ")

                        if number_code[len(number_code)-1].isalpha() == True:
                            number_code.pop(len(number_code)-1)

                        number_code = number_code[len(number_code)-1]
                        vessel_code = main_ws['B{}'.format(j)].value
                        
                        data = {
                            "MTR": {
                                "VAN": 0,
                                "PRR": 0, 
                                "HAL": 0,
                                "RF": 0,
                                'MTR': 0,
                            },
                            "TOR": {
                                "VAN": 0, 
                                "PRR": 0,
                                "HAL": 0,
                                "RF": 0,
                                "MTR": 0,  
                            },
                            "VAN": {
                                "VAN": 0,
                                "PRR": 0,
                                "HAL": 0,
                                "RF": 0,
                                "MTR": 0,
                            },
                        }

                        #data[Booking Office][Port of loading]

                        cell_range = read_ws['F2':'F{}'.format(read_ws.max_row)]
                        index = 2

                        target_code = "{}-{}-{}".format(shipping_lane_code,vessel_code,number_code) 

                        for i in cell_range:
                            export_code = i[0].value[:-1].strip()

                            if export_code == target_code:                        
                                Booking_Office_Code = read_ws['A{}'.format(index)].value
                                
                                Pol =  read_ws['E{}'.format(index)].value
                                
                                current_teu = read_ws['G{}'.format(index)].value
                                rifer_value = read_ws['L{}'.format(index)].value

                                if Pol == "MTR" or Pol == "HAL":
                                    data[Booking_Office_Code]["VAN"] = data[Booking_Office_Code]["VAN"] + current_teu
                                else:
                                    data[Booking_Office_Code][Pol] = data[Booking_Office_Code][Pol] + current_teu

                                data[Booking_Office_Code]["RF"] = data[Booking_Office_Code]["RF"] + rifer_value

                            if index != read_ws.max_row:
                                index += 1

                        #Vancouver Office
                        main_ws['L{}'.format(j)] = data['VAN']['PRR'] 
                        main_ws['N{}'.format(j)] = data['VAN']['VAN']
                        main_ws['P{}'.format(j)] = data['VAN']['RF']

                        #Montreal Office
                        main_ws['Q{}'.format(j)] = data['MTR']['PRR']
                        main_ws['R{}'.format(j)] = data['MTR']['VAN']
                        main_ws['S{}'.format(j)] = data['MTR']['RF']

                        #Toronto Office
                        main_ws['T{}'.format(j)] = data['TOR']['PRR']
                        main_ws['U{}'.format(j)] = data['TOR']['VAN']
                        main_ws['V{}'.format(j)] = data['TOR']['RF']

        main_workbook.save("static/Sheets/{}".format(write_file.filename))
    try: 
        return send_from_directory(current_app.config["SHEET_STORAGE"], write_file.filename, as_attachment=True)
    except FileNotFoundError:
        print()
        os.abort(404)
        



